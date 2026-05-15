import os, textwrap, math
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import fitz
from PIL import Image, ImageDraw, ImageFont

BASE='/root/Hermes-Agent'
XLSX_IN='/root/.hermes/cache/documents/doc_f6c73634a7da_8. EW_E&W_B1T1_Assmt.xlsx'
PDF_IN='/root/.hermes/cache/documents/doc_d37e65ca46e5_4. EW_E&W_B1T1_RM_A2.pdf'
XLSX_OUT=f'{BASE}/8. EW_E&W_B1T1_Assmt_Marathi.xlsx'
PDF_OUT=f'{BASE}/4. EW_E&W_B1T1_RM_A2_Marathi.pdf'
FONT='/usr/share/fonts/truetype/noto/NotoSansDevanagari-Regular.ttf'
FONT_BOLD='/usr/share/fonts/truetype/noto/NotoSansDevanagari-Bold.ttf'
FONT_LATIN='/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf'
FONT_LATIN_BOLD='/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf'
FONT_ITALIC='/usr/share/fonts/truetype/noto/NotoSansDevanagari-Regular.ttf'

# ---------------- XLSX ----------------
translations = {
    'Question No':'प्रश्न क्रमांक',
    'Question Text':'प्रश्न मजकूर',
    'Option Count':'पर्याय संख्या',
    'Correct Option':'योग्य पर्याय',
    'Option 1':'पर्याय १',
    'Option 2':'पर्याय २',
    'Option 3':'पर्याय ३',
    'Option 4':'पर्याय ४',
    'A motor is running. A smart system sends an alert: “Motor temperature is slowly increasing.” The motor is still running.\nWhat does this alert help the electrician do?\nSelect the correct option.': 'एक मोटर चालू आहे. स्मार्ट प्रणाली सूचना पाठवते: “मोटरचे तापमान हळूहळू वाढत आहे.” मोटर अजूनही चालू आहे.\nही सूचना इलेक्ट्रिशियनला काय करण्यास मदत करते?\nयोग्य पर्याय निवडा.',
    'Wait until the motor stops.':'मोटर बंद होईपर्यंत थांबा.',
    'Check the motor early and stop a problem.':'मोटर लवकर तपासा आणि समस्या थांबवा.',
    'Close the factory.':'कारखाना बंद करा.',
    'Ignore the message.':'संदेशाकडे दुर्लक्ष करा.',
    'In an old electrical system, the machine does not give any warning before a fault. The electrician knows about a fault only after the machine stops.\nWhat does this lead to?\nSelect the correct option.':'जुन्या विद्युत प्रणालीमध्ये, बिघाड होण्यापूर्वी मशीन कोणतीही सूचना देत नाही. मशीन बंद झाल्यानंतरच इलेक्ट्रिशियनला बिघाडाबद्दल कळते.\nयामुळे काय होते?\nयोग्य पर्याय निवडा.',
    'Work stops suddenly and people must wait for the machine to be repaired.':'काम अचानक थांबते आणि लोकांना मशीन दुरुस्त होईपर्यंत थांबावे लागते.',
    'The machine stops, but the work continues without any delay.':'मशीन थांबते, पण काम कोणत्याही विलंबाशिवाय सुरू राहते.',
    'The machine stops, and the fault gets corrected automatically by itself.':'मशीन थांबते आणि बिघाड आपोआप दुरुस्त होतो.',
    'Enough time for the electrician to plan the repair.':'इलेक्ट्रिशियनला दुरुस्तीचे नियोजन करण्यासाठी पुरेसा वेळ मिळतो.',
    'Compare these two factories:\n• Factory A waits for machines to fail\n• Factory B uses early warnings and plans maintainence\nWhich workplace is safer to work in and why?\nSelect the correct option.':'या दोन कारखान्यांची तुलना करा:\n• कारखाना A मशीन बिघडेपर्यंत थांबतो\n• कारखाना B आधीच सूचना वापरतो आणि देखभालीचे नियोजन करतो\nकाम करण्यासाठी कोणते कार्यस्थळ अधिक सुरक्षित आहे आणि का?\nयोग्य पर्याय निवडा.',
    'Factory A, because faults are found only after the machine stops.':'कारखाना A, कारण मशीन थांबल्यानंतरच बिघाड आढळतात.',
    'Factory B, because early warnings help fix problems before they become dangerous.':'कारखाना B, कारण आधीच मिळालेल्या सूचना समस्या धोकादायक होण्यापूर्वी दुरुस्त करण्यास मदत करतात.',
    'Factory A, because repairs are done only when there is a breakdown.':'कारखाना A, कारण बिघाड झाल्यावरच दुरुस्ती केली जाते.',
    'Factory B, because machines will never fail in this factory.':'कारखाना B, कारण या कारखान्यात मशीन कधीही बिघडणार नाहीत.',
    'Suppose you get a warning early and fix a small problem. What will happen next?\nSelect the correct option.':'समजा तुम्हाला आधीच सूचना मिळाली आणि तुम्ही छोटी समस्या दुरुस्त केली. पुढे काय होईल?\nयोग्य पर्याय निवडा.',
    'The machine will break suddenly.':'मशीन अचानक बिघडेल.',
    'Work will stop for a long time.':'काम बराच वेळ थांबेल.',
    'Big problems can be prevented.':'मोठ्या समस्या टाळता येतील.',
    'Nothing will change.':'काहीही बदलणार नाही.',
    'A smart system says “Overload alert.” What is the best thing to do?\nSelect the correct option.':'स्मार्ट प्रणाली सांगते: “ओव्हरलोड सूचना.” करण्यासाठी सर्वोत्तम गोष्ट कोणती?\nयोग्य पर्याय निवडा.',
    'Wait for the system to shut down on its own.':'प्रणाली स्वतः बंद होईपर्यंत थांबा.',
    'Check the issue and fix it immediately.':'समस्या तपासा आणि लगेच दुरुस्त करा.',
    'Ignore the alert.':'सूचनेकडे दुर्लक्ष करा.',
    'Turn off all other systems to wait for the alert to go away.':'सूचना नाहीशी होईपर्यंत इतर सर्व प्रणाली बंद करा.',
}

wb=load_workbook(XLSX_IN)
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value in translations:
                c.value=translations[c.value]
                c.font=copy(c.font)
                c.font=Font(name='Noto Sans Devanagari', size=c.font.sz or 11, bold=c.font.bold, italic=c.font.italic, color=c.font.color)
                c.alignment=copy(c.alignment)
                c.alignment=Alignment(wrap_text=True, vertical=c.alignment.vertical or 'top', horizontal=c.alignment.horizontal)
wb.save(XLSX_OUT)

# ---------------- PDF helpers ----------------
def pil_font(size, bold=False, latin=False):
    return ImageFont.truetype(FONT_BOLD if bold else FONT, size)

def wrap_text(draw, text, font, max_w):
    lines=[]
    for para in text.split('\n'):
        words=para.split(' ')
        line=''
        for w in words:
            test=(line+' '+w).strip()
            if draw.textbbox((0,0), test, font=font)[2] <= max_w or not line:
                line=test
            else:
                lines.append(line); line=w
        lines.append(line)
    return lines

def insert_text_image(page, rect, text, size=12, color=(0,0,0), bold=False, align='left', line_spacing=1.35, bg=None):
    scale=3
    w=max(1,int(rect.width*scale)); h=max(1,int(rect.height*scale))
    img=Image.new('RGBA',(w,h),(255,255,255,0) if bg is None else bg)
    d=ImageDraw.Draw(img)
    font=pil_font(int(size*scale), bold=bold)
    lines=wrap_text(d, text, font, w-6*scale)
    y=0
    for line in lines:
        bbox=d.textbbox((0,0), line, font=font)
        tw=bbox[2]-bbox[0]
        x=0
        if align=='center': x=(w-tw)//2
        elif align=='right': x=w-tw
        d.text((x,y), line, font=font, fill=color+(255,))
        y += int(size*scale*line_spacing)
        if y>h: break
    tmp=f'/tmp/text_{os.getpid()}_{abs(hash((text,rect.x0,rect.y0)))%10**9}.png'
    img.save(tmp)
    page.insert_image(rect, filename=tmp, overlay=True)
    try: os.remove(tmp)
    except: pass

def cover(page, x0,y0,x1,y1, color=(1,1,1)):
    page.draw_rect(fitz.Rect(x0,y0,x1,y1), color=color, fill=color, overlay=True)

def bullet_text(page, x, y, text, w=430, size=10.5, color=(0,0,0), bullet=True):
    if bullet:
        page.draw_circle(fitz.Point(x+4,y+6), 2.2, color=(0,0,0), fill=(0,0,0), overlay=True)
        insert_text_image(page, fitz.Rect(x+18,y,w+x,y+28), text, size=size)
    else:
        insert_text_image(page, fitz.Rect(x,y,w+x,y+28), text, size=size)

# Marathi page content
P={
1:{
'title':'स्मार्ट तंत्रज्ञानाच्या मदतीने स्मार्ट कंपन्या दिवे कसे चालू ठेवतात?',
'sub':'कामाच्या ठिकाणची खरी परिस्थिती!',
'paras':[
'कारखान्यातील एक सामान्य दिवस कल्पना करा. मशीन सुरळीत चालू आहेत. कामगार आपापले काम करण्यात व्यस्त आहेत. दिवे चालू आहेत आणि सर्व काही व्यवस्थित चालू आहे. अचानक—क्लिक! वीज जाते. दिवे बंद होतात, मशीन थांबतात आणि कारखाना शांत होतो. कामगार काम थांबवून वाट पाहतात. पर्यवेक्षक देखभाल पथकाला बोलावतात. कामाला उशीर होतो.',
'अगदी थोडा वेळ वीज गेली तरी अनेक समस्या निर्माण होऊ शकतात:',
],
'bullets':['उत्पादनाचे नुकसान — काम पूर्णपणे थांबते','साहित्याचे नुकसान — अर्धवट तयार वस्तू खराब होऊ शकतात किंवा मशीनमध्ये अडकू शकतात','अतिरिक्त दुरुस्ती वेळ — कामगारांना समस्या शोधून दुरुस्त करावी लागते','सुरक्षेचा धोका — मशीन अचानक थांबल्यास धोका निर्माण होऊ शकतो'],
'after':'अशा वीज समस्या फक्त कारखान्यांतच नाही, तर कार्यालये, रुग्णालये, शॉपिंग मॉल आणि डेटा सेंटर्समध्येही होऊ शकतात. डेटा सेंटर्समध्ये वेबसाइट्स आणि इंटरनेट सेवा चालवणारे संगणक असतात. काही मिनिटे वीज नसली तरी खूप मोठे नुकसान होऊ शकते.',
'h2':'वीज बिघाडामुळे खूप पैसा का खर्च होतो?',
'p2':'अनेक उद्योगांमध्ये मशीन सतत चालू असतात.\nवीज अचानक गेली तर:',
'bullets2':['मशीन पुन्हा सुरू होण्यासाठी वेळ लागतो','गुणवत्ता तपासणी पुन्हा करावी लागते','उत्पादन लक्ष्य चुकतात'],
'end':'यामुळे कंपन्यांचा वेळ आणि पैसा दोन्ही वाया जातात. काच आणि इलेक्ट्रॉनिक्ससारख्या उच्च-मूल्य उद्योगांना वीज बिघाडानंतर दीर्घ बंद आणि महागड्या दुरुस्तींमुळे मोठे नुकसान सहन करावे लागते.',
'call':'भारतातील अनेक कारखान्यांमध्ये, नियोजनाशिवाय झालेला १०–१५ मिनिटांचा वीज बिघाडही तासन्तास उत्पादन नुकसान करू शकतो, कारण मशीन पुन्हा सुरू करण्यापूर्वी थंड करणे, पुनर्मापन आणि सुरक्षा तपासणी करावी लागते.'
},
2:{
'h':'समस्या लवकर थांबवण्यासाठी स्मार्ट तंत्रज्ञानाचा वापर',
'p':'आज अनेक ठिकाणी विद्युत पॅनेल आणि मशीनची नियमित पाहणी केली जाते. यामुळे बिघाड होण्यापूर्वी कामगारांना आधीच सूचना मिळते. खालील दोन पद्धतींची तुलना करा.',
'cap':'तक्ता १: जुनी आणि स्मार्ट पद्धतींची तुलना',
'headers':['जुनी पद्धत (पारंपरिक पद्धत)','स्मार्ट पद्धत (स्मार्ट तंत्रज्ञान)'],
'rows':[['मशीन आधी बिघडते','स्मार्ट उपकरणे मशीन सतत तपासत राहतात'],['काम थांबते','प्रणाली आधीच सूचना देते'],['मग कामगार समस्या तपासून दुरुस्त करतात','मशीन बिघडण्यापूर्वीच कामगार दुरुस्ती करतात'],['यामुळे विलंब आणि त्रास होतो','स्मार्ट पद्धत वेळ, पैसा वाचवते आणि सर्वांना सुरक्षित ठेवते']],
'intro':'बिघाड टाळण्यासाठी अनेक कंपन्या स्मार्ट प्रणाली वापरतात. यामुळे समस्या लवकर शोधण्यास मदत होते. काही उदाहरणे खाली दिली आहेत:',
'boxes':[
('आयओटी (इंटरनेट ऑफ थिंग्स)','मशीनवर किंवा पॅनेलमध्ये बसवलेले सेन्सर/मीटर व्होल्टेज, करंट, लोड आणि तापमान यांसारख्या रीडिंग्स सतत तपासतात. ही रीडिंग्स मॉनिटरिंग प्रणालीकडे पाठवली जातात, जी आधीच सूचना देऊ शकते. (काही मशीन कंपन आणि ट्रिप स्थितीही तपासतात.)'),
('एआय (कृत्रिम बुद्धिमत्ता)','एआय ही अशी प्रणाली आहे जी मशीन/पॅनेलच्या रीडिंग्सकडे पाहते आणि असामान्य नमुने शोधते. ती अशा प्रश्नांची उत्तरे शोधण्यास मदत करते:\n“ही रीडिंग या मशीनसाठी सामान्य आहे का?”\n“काहीतरी हळूहळू बिघडत आहे का?”\n“हा बिघाडाचा सुरुवातीचा संकेत आहे का?”'),
('प्रेडिक्टिव्ह अॅनालिसिस','याचा अर्थ सेन्सर/मीटरच्या रीडिंग्स वापरून समस्या आधीच ओळखणे. प्रणाली बिघाड होण्यापूर्वी सूचना देते, उदाहरणार्थ:\n“तापमान वाढत आहे.”\n“लोड/करंट सामान्यपेक्षा जास्त आहे.”\n“हा भाग तपासला नाही तर लवकरच बिघडू शकतो.”')]
},
3:{
'top':'ही साधने वापरून कंपन्या मशीन बिघडण्याची वाट पाहत नाहीत. त्या बिघाड गंभीर होण्यापूर्वी तपासणी आणि कृती करण्यासाठी वेळ मिळवतात.',
'call':'हे जणू त्रास येण्यापूर्वीच सावध करणारा सुपरहिरो असल्यासारखे आहे.\nसोप्या भाषेत: बिघाडाची वाट पाहू नका. लवकर कृती करा.',
'h':'कामाच्या ठिकाणच्या खऱ्या गोष्टी: स्मार्ट तंत्रज्ञान कृतीत',
'intro':'आता कामाच्या ठिकाणच्या काही छोट्या गोष्टी वाचूया. या गोष्टी दाखवतात की सेन्सर आणि स्मार्ट प्रणाली आधीच सूचना देतात, त्यामुळे बिघाड होण्यापूर्वी दोष दुरुस्त करता येतात.',
'h1':'कंपनी ए — रासायनिक उत्पादन कंपनी',
'p1':'कंपनी ए रसायने तयार करते, त्यामुळे अनेक मशीन सतत चालू राहणे आवश्यक असते. वीज थांबली तर उत्पादन बिघडू शकते आणि साहित्य वाया जाऊ शकते. म्हणून प्लांटने ट्रान्सफॉर्मर आणि मुख्य पॅनेलसारख्या महत्त्वाच्या विद्युत उपकरणांवर मॉनिटरिंग सेन्सर/मीटर बसवले. ही उपकरणे तापमान आणि लोड करंट यांसारखी रीडिंग्स सतत तपासत राहतात. एके दिवशी मॉनिटरिंग प्रणालीने सूचना पाठवली: “ट्रान्सफॉर्मरचे तापमान वाढत आहे / लोड सामान्यपेक्षा जास्त आहे.” देखभाल पथक साइटवर पोहोचले, सुरक्षा पायऱ्या पाळल्या आणि ट्रान्सफॉर्मर व पॅनेल तपासले. त्यांनी समस्या लवकर शोधली (उदा. जास्त लोडमुळे ओव्हरहिटिंग किंवा सप्लाय लाइनमधील सैल जोडणी) आणि ती दुरुस्त केली.',
'res1':'निकाल: ट्रान्सफॉर्मर बिघडला नाही, शटडाउन झाले नाही आणि प्लांट सुरक्षितपणे चालू राहिला.',
'h2':'अन्न प्रक्रिया प्लांटमध्ये',
'p2':'अन्न प्रक्रिया प्लांटमध्ये मशीन वेळेवर चालू राहणे आवश्यक असते. वीज थांबली तर संपूर्ण उत्पादन रेषेला उशीर होऊ शकतो. म्हणून कारखाना ट्रान्सफॉर्मर आणि मुख्य पॅनेलसारख्या महत्त्वाच्या विद्युत उपकरणांसाठी मॉनिटरिंग प्रणाली वापरतो. सेन्सर/मीटर तापमान, करंट (अँप्स) आणि व्होल्टेज यांसारखी रीडिंग्स तपासतात आणि स्क्रीन/मोबाइलवर सूचना दाखवतात. एके दिवशी प्रणालीने सूचना दिली: “ट्रान्सफॉर्मरचे तापमान वाढत आहे / करंटमध्ये असंतुलन दिसत आहे—कृपया तपासा.” दीपकचे देखभाल पथक साइटवर गेले आणि सुरक्षा पायऱ्या पाळल्या. त्यांनी ट्रान्सफॉर्मर क्षेत्र आणि पॅनेल जोडण्या तपासल्या. त्यांनी समस्या लवकर शोधली (टर्मिनलवर सैल जोडणी/हीटिंग किंवा एका फीडरवर ओव्हरलोड) आणि ती दुरुस्त केली. हे दुर्लक्षित केले असते तर ट्रान्सफॉर्मर ट्रिप किंवा बिघडू शकला असता, ज्यामुळे अनेक तासांचा डाउनटाइम झाला असता. त्यांनी लवकर कृती केल्यामुळे उत्पादन शटडाउनशिवाय सुरू राहिले.'
},
4:{
'h':'एक्सवायझेड प्रायव्हेट लिमिटेड — कार्यालयीन इमारती',
'p':'एक्सवायझेड प्रायव्हेट लिमिटेड भारतातील अनेक कार्यालयीन इमारती व्यवस्थापित करते. या इमारतींमध्ये त्यांना हे करायचे होते:',
'bul':['वीज वाया जाणे कमी करणे, आणि','ओव्हरलोड व वारंवार ट्रिप होणे अशा वीज समस्या टाळणे.'],
'p2':'म्हणून त्यांनी केंद्रीय मॉनिटरिंग रूम (ऊर्जा कमांड सेंटर) तयार केली. इमारतींमधील सेन्सर/मीटर ऊर्जा वापर, करंट, व्होल्टेज आणि लोड यांसारखी रीडिंग्स प्रणालीकडे पाठवतात.',
'p3':'प्रणालीने त्यांना अशा गोष्टी लक्षात आणून दिल्या:',
'bul2':['कार्यालयीन वेळेनंतरही काही मजले जास्त वीज वापरत होते','विशिष्ट वेळी अचानक जास्त लोड येत होता','ओव्हरहिटिंग किंवा ट्रिप होण्यास कारणीभूत असामान्य बदल'],
'p4':'त्यानंतर सुविधा पथकाने अनावश्यक लोड बंद करणे, लोड संतुलित करणे आणि हीटिंग किंवा वारंवार ट्रिप दिसलेल्या पॅनेलची तपासणी करणे अशी कृती केली.',
'res':'निकाल: इमारतींनी वीज अधिक कार्यक्षमतेने वापरली आणि अनपेक्षित वीज समस्या कमी झाल्या.',
'learn':'आपण काय शिकलो: मॉनिटरिंगमुळे पथकांना समस्या लवकर दिसतात आणि त्या मोठ्या बिघाडात बदलण्यापूर्वी कृती करता येते.',
'key':'या गोष्टींमधून मुख्य शिकवण',
'items':['वीज समस्या खूप खर्चिक असू शकतात. अगदी थोडा आउटेजही मशीन थांबवू शकतो, उत्पादने खराब करू शकतो आणि पैसा खर्च करू शकतो','स्मार्ट तंत्रज्ञान समस्या टाळण्यास मदत करते. आयओटी सेन्सर मशीनवर लक्ष ठेवतात, एआय डेटा अभ्यासते आणि प्रेडिक्टिव्ह अॅनालिटिक्स आधीच सूचना देते','समस्या लवकर दुरुस्त केल्याने पैसा आणि वेळ वाचतो. मशीन बिघडेपर्यंत थांबणे खर्चिक असते','स्मार्ट व्यवस्थापन पृथ्वीला मदत करते. ऊर्जा वापर कमी होतो, मशीन जास्त काळ टिकतात आणि नवीकरणीय ऊर्जा अधिक चांगली वापरता येते','सुरक्षा सुधारते. कामगारांना अचानक बिघाडांना सामोरे जावे लागत नाही, त्यामुळे कारखाने अधिक सुरक्षित होतात']
},
5:{
'h':'मोठा संदेश',
'p':'स्मार्ट संस्था समस्या येण्याची वाट पाहत नाहीत. त्या पाहतात, शिकतात आणि अंदाज घेतात.',
'p2':'एआय, आयओटी आणि प्रेडिक्टिव्ह अॅनालिटिक्सच्या मदतीने त्या करू शकतात:',
'bul':['मशीन चालू ठेवणे','वीज बिघाड टाळणे','कामगारांचे संरक्षण करणे','पैसा वाचवणे','पर्यावरणाची काळजी घेणे'],
'call':'हे जणू सतत मशीनवर लक्ष ठेवणारा स्मार्ट मदतनीस असल्यासारखे आहे.\nम्हणूनच संस्था दिवे चालू ठेवतात आणि भविष्याकरिता तयार राहतात.',
'source':'स्रोत श्रेय: हे शिक्षण साहित्य दिप्ती पांडा यांनी प्रकाशित केलेल्या “शाश्वत भविष्यासाठी तुमचा मार्ग शोधा: सक्रिय विद्युत मालमत्ता व्यवस्थापनाद्वारे अनियोजित डाउनटाइम कमी करणे” या लेखावर आधारित तयार केले आहे. मूळ लेखाचा संदर्भ मूळ इंग्रजी दस्तऐवजात दिला आहे.'
}
}

# Build PDF by overlaying text areas on original template
src=fitz.open(PDF_IN)
for i,page in enumerate(src):
    # cover content area, preserve header/footer/logos/page number
    cover(page, 65, 86, 548, 740)
    blue=(0,58,100); darkblue=(0,70,100); orange=(1,0.5,0)
    n=i+1
    if n==1:
        d=P[1]
        insert_text_image(page, fitz.Rect(95,92,520,145), d['title'], 15, color=(47,85,130), bold=True, align='center')
        insert_text_image(page, fitz.Rect(76,160,470,182), d['sub'], 12, color=darkblue, bold=True)
        insert_text_image(page, fitz.Rect(76,192,538,264), d['paras'][0], 9.7)
        insert_text_image(page, fitz.Rect(76,275,510,292), d['paras'][1], 10)
        y=302
        for b in d['bullets']:
            bullet_text(page, 85, y, b, w=455, size=9.3); y+=21
        insert_text_image(page, fitz.Rect(76,386,538,440), d['after'], 9.2)
        insert_text_image(page, fitz.Rect(76,450,510,472), d['h2'], 12, color=darkblue, bold=True)
        insert_text_image(page, fitz.Rect(76,484,520,524), d['p2'], 9.5)
        y=536
        for b in d['bullets2']:
            bullet_text(page, 85, y, b, w=455, size=9.3); y+=20
        insert_text_image(page, fitz.Rect(76,600,538,638), d['end'], 9.1)
        # callout approximation
        page.draw_rect(fitz.Rect(126,646,548,722), color=(0,0.45,0.8), fill=(0.9,0.97,1), width=1.5, overlay=True)
        insert_text_image(page, fitz.Rect(150,665,528,710), d['call'], 8.5, color=(0,0,0))
    elif n==2:
        d=P[2]
        insert_text_image(page, fitz.Rect(76,94,520,116), d['h'], 12.5, color=darkblue, bold=True)
        insert_text_image(page, fitz.Rect(76,126,538,160), d['p'], 9.4)
        insert_text_image(page, fitz.Rect(200,170,410,184), d['cap'], 8.2, color=(0,0,0), align='center')
        # table
        x0,y0,x1,y1=76,186,538,286; rowh=20; colmid=275
        page.draw_rect(fitz.Rect(x0,y0,x1,y1), color=(0,0,0), fill=None, overlay=True)
        page.draw_rect(fitz.Rect(x0,y0,x1,y0+rowh), color=(0.12,0.38,0.62), fill=(0.12,0.38,0.62), overlay=True)
        page.draw_line(fitz.Point(colmid,y0),fitz.Point(colmid,y1),color=(0,0,0),overlay=True)
        for rr in range(1,5): page.draw_line(fitz.Point(x0,y0+rowh*rr),fitz.Point(x1,y0+rowh*rr),color=(0,0,0),overlay=True)
        insert_text_image(page, fitz.Rect(x0+5,y0+4,colmid-4,y0+20), d['headers'][0], 8.5, color=(255,255,255), bold=True)
        insert_text_image(page, fitz.Rect(colmid+5,y0+4,x1-4,y0+20), d['headers'][1], 8.5, color=(255,255,255), bold=True)
        for r,row in enumerate(d['rows']):
            y=y0+rowh*(r+1)+3
            insert_text_image(page, fitz.Rect(x0+5,y,colmid-4,y+18), row[0], 7.8)
            insert_text_image(page, fitz.Rect(colmid+5,y,x1-4,y+18), row[1], 7.8)
        insert_text_image(page, fitz.Rect(76,312,538,346), d['intro'], 9.5)
        y=360
        heights=[96,122,112]
        for (title,body),h in zip(d['boxes'],heights):
            page.draw_rect(fitz.Rect(86,y,538,y+h), color=(1,0.45,0), fill=(1,0.98,0.95), width=1.3, overlay=True)
            insert_text_image(page, fitz.Rect(150,y+12,525,y+30), title, 9.8, bold=True)
            insert_text_image(page, fitz.Rect(150,y+34,525,y+h-8), body, 8.2)
            y+=h+15
    elif n==3:
        d=P[3]
        insert_text_image(page, fitz.Rect(76,94,538,132), d['top'], 9.4)
        page.draw_rect(fitz.Rect(92,145,538,210), color=(0,0.72,0.9), fill=(0.88,0.98,1), width=1.5, overlay=True)
        insert_text_image(page, fitz.Rect(155,160,520,195), d['call'], 8.8, color=darkblue, align='center')
        insert_text_image(page, fitz.Rect(76,229,538,252), d['h'], 12, color=darkblue, bold=True)
        insert_text_image(page, fitz.Rect(76,262,538,296), d['intro'], 9.4)
        insert_text_image(page, fitz.Rect(76,307,538,326), d['h1'], 9.7, bold=True)
        insert_text_image(page, fitz.Rect(76,331,538,464), d['p1'], 8.7)
        insert_text_image(page, fitz.Rect(76,470,538,492), d['res1'], 9.0)
        insert_text_image(page, fitz.Rect(76,500,538,518), d['h2'], 9.7, bold=True)
        insert_text_image(page, fitz.Rect(76,522,538,710), d['p2'], 8.4)
    elif n==4:
        d=P[4]
        insert_text_image(page, fitz.Rect(76,94,538,112), d['h'], 9.8, bold=True)
        insert_text_image(page, fitz.Rect(76,117,538,137), d['p'], 9.0)
        y=143
        for b in d['bul']: bullet_text(page,100,y,b,w=430,size=8.8); y+=22
        insert_text_image(page, fitz.Rect(76,188,538,222), d['p2'], 8.9)
        insert_text_image(page, fitz.Rect(76,234,538,252), d['p3'], 9.0)
        y=260
        for b in d['bul2']: bullet_text(page,100,y,b,w=430,size=8.7); y+=21
        insert_text_image(page, fitz.Rect(76,324,538,360), d['p4'], 8.9)
        insert_text_image(page, fitz.Rect(76,370,538,388), d['res'], 8.9)
        insert_text_image(page, fitz.Rect(76,400,538,420), d['learn'], 8.8, color=darkblue, bold=True)
        insert_text_image(page, fitz.Rect(76,430,538,452), d['key'], 9.0)
        colors=[(0.3,0.8,0.95),(0.1,0.75,0.65),(0.45,0.55,0.58),(0.95,0.72,0.1),(0.95,0.25,0.2)]
        y=468
        for idx,item in enumerate(d['items'],1):
            col=colors[idx-1]
            page.draw_rect(fitz.Rect(125,y-5,538,y+34), color=col, fill=(1,1,1), width=1.0, overlay=True)
            page.draw_circle(fitz.Point(120,y+14), 11, color=col, fill=col, overlay=True)
            insert_text_image(page, fitz.Rect(113,y+6,128,y+22), ['०१','०२','०३','०४','०५'][idx-1], 6.2, color=(255,255,255), bold=True, align='center')
            insert_text_image(page, fitz.Rect(140,y+1,530,y+31), item, 7.6)
            y += 43
    elif n==5:
        d=P[5]
        insert_text_image(page, fitz.Rect(76,94,538,116), d['h'], 12, color=darkblue, bold=True)
        insert_text_image(page, fitz.Rect(76,126,538,145), d['p'], 9.2)
        insert_text_image(page, fitz.Rect(76,153,538,172), d['p2'], 9.2)
        y=180
        for b in d['bul']: bullet_text(page,85,y,b,w=300,size=9.1); y+=20
        page.draw_rect(fitz.Rect(315,180,538,275), color=(0,0.72,0.9), fill=(1,1,1), width=1.3, overlay=True)
        insert_text_image(page, fitz.Rect(340,200,520,260), d['call'], 8.4, align='center')
        page.draw_rect(fitz.Rect(76,310,538,420), color=(0,0,0), fill=(1,1,1), width=0.8, overlay=True)
        insert_text_image(page, fitz.Rect(82,316,532,410), d['source'], 8.0, color=(0,0,0))

src.save(PDF_OUT, garbage=4, deflate=True)
print('WROTE', XLSX_OUT)
print('WROTE', PDF_OUT)
